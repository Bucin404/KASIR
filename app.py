from datetime import datetime, timedelta
from functools import wraps
import os
import json
import qrcode
from io import BytesIO
import base64

from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename

from config import config
from models import db, User, Role, Permission, Category, MenuItem, Table, Order, OrderItem, Payment, Income, Setting, Cart, CartItem

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(config['development'])

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Custom Jinja2 filters
def format_number_filter(value):
    """Format angka dengan pemisah ribuan"""
    try:
        return f"{int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return value

def format_currency(value):
    """Format sebagai mata uang Rupiah"""
    try:
        return f"Rp {int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['format_number'] = format_number_filter
app.jinja_env.filters['format_currency'] = format_currency

# Context processor to make config available in all templates
@app.context_processor
def inject_config():
    return {
        'config': {
            'MIDTRANS_CLIENT_KEY': app.config.get('MIDTRANS_CLIENT_KEY', 'SB-Mid-client-XXXXXX'),
            'MIDTRANS_IS_PRODUCTION': app.config.get('MIDTRANS_IS_PRODUCTION', False),
            'APP_NAME': 'Dapoer Teras Obor'
        }
    }

# Permission decorator
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_permission(permission):
                flash('Anda tidak memiliki akses ke halaman ini.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not any(current_user.has_role(role) for role in roles):
                flash('Anda tidak memiliki akses ke halaman ini.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Initialize database and seed data
def run_migrations():
    """Run database migrations for existing databases"""
    from sqlalchemy import inspect, text
    
    inspector = inspect(db.engine)
    
    # Check if payments table exists and add snap_token column if missing
    if 'payments' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('payments')]
        if 'snap_token' not in columns:
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE payments ADD COLUMN snap_token VARCHAR(255)'))
                conn.commit()
            print("Added snap_token column to payments table")
    
    # Check if cart table exists (new feature)
    if 'cart' not in inspector.get_table_names():
        # db.create_all will handle this
        pass
    
    if 'cart_item' not in inspector.get_table_names():
        # db.create_all will handle this
        pass

def init_db():
    with app.app_context():
        db.create_all()
        
        # Run migrations for existing databases
        run_migrations()
        
        # Create default permissions
        permissions_data = [
            ('view_dashboard', 'Dapat melihat dashboard'),
            ('manage_orders', 'Dapat mengelola pesanan'),
            ('manage_menu', 'Dapat mengelola menu'),
            ('manage_users', 'Dapat mengelola pengguna'),
            ('manage_tables', 'Dapat mengelola meja'),
            ('view_reports', 'Dapat melihat laporan'),
            ('manage_settings', 'Dapat mengelola pengaturan'),
            ('process_payment', 'Dapat memproses pembayaran'),
            ('view_income', 'Dapat melihat penghasilan'),
            ('manage_income', 'Dapat mengelola penghasilan'),
        ]
        
        for perm_name, perm_desc in permissions_data:
            if not Permission.query.filter_by(name=perm_name).first():
                perm = Permission(name=perm_name, description=perm_desc)
                db.session.add(perm)
        
        db.session.commit()
        
        # Create default roles
        roles_data = {
            'admin': {
                'description': 'Administrator dengan akses penuh',
                'permissions': [p[0] for p in permissions_data]
            },
            'manager': {
                'description': 'Manager dengan akses laporan dan manajemen',
                'permissions': ['view_dashboard', 'manage_orders', 'manage_menu', 'view_reports', 'manage_tables', 'process_payment', 'view_income']
            },
            'kasir': {
                'description': 'Kasir untuk proses pembayaran',
                'permissions': ['view_dashboard', 'manage_orders', 'process_payment']
            },
            'customer': {
                'description': 'Pelanggan untuk memesan online',
                'permissions': ['view_dashboard']
            }
        }
        
        for role_name, role_data in roles_data.items():
            role = Role.query.filter_by(name=role_name).first()
            if not role:
                role = Role(name=role_name, description=role_data['description'])
                db.session.add(role)
                db.session.commit()
            
            # Add permissions to role
            for perm_name in role_data['permissions']:
                perm = Permission.query.filter_by(name=perm_name).first()
                if perm and perm not in role.permissions:
                    role.permissions.append(perm)
        
        db.session.commit()
        
        # Create default admin user
        if not User.query.filter_by(username='admin').first():
            admin_role = Role.query.filter_by(name='admin').first()
            admin = User(
                username='admin',
                email='admin@kasir.com',
                full_name='Administrator'
            )
            admin.set_password('admin123')
            admin.roles.append(admin_role)
            db.session.add(admin)
        
        # Create default kasir user
        if not User.query.filter_by(username='kasir').first():
            kasir_role = Role.query.filter_by(name='kasir').first()
            kasir = User(
                username='kasir',
                email='kasir@kasir.com',
                full_name='Kasir Utama'
            )
            kasir.set_password('kasir123')
            kasir.roles.append(kasir_role)
            db.session.add(kasir)
        
        db.session.commit()
        
        # Create default categories
        categories_data = [
            ('Nasi Goreng', 'Menu nasi goreng berbagai varian', 'fa-bowl-rice', 1),
            ('Mie', 'Menu mie berbagai varian', 'fa-utensils', 2),
            ('Kwetiau', 'Menu kwetiau berbagai varian', 'fa-plate-wheat', 3),
            ('Menu Lain', 'Menu lainnya', 'fa-drumstick-bite', 4),
            ('Paket', 'Menu paket hemat', 'fa-box', 5),
            ('Snack', 'Makanan ringan', 'fa-cookie', 6),
            ('Minuman', 'Berbagai minuman segar', 'fa-mug-hot', 7),
        ]
        
        for cat_name, cat_desc, cat_icon, cat_order in categories_data:
            if not Category.query.filter_by(name=cat_name).first():
                cat = Category(name=cat_name, description=cat_desc, icon=cat_icon, order=cat_order)
                db.session.add(cat)
        
        db.session.commit()
        
        # Create menu items from PDF menu (Solaria style)
        seed_menu_items()
        
        # Create default tables
        for i in range(1, 21):
            table_num = f"{i:02d}"
            if not Table.query.filter_by(number=table_num).first():
                table = Table(
                    number=table_num,
                    name=f"Meja {i}",
                    capacity=4 if i <= 15 else 6
                )
                db.session.add(table)
        
        db.session.commit()
        print("Database initialized successfully!")

def seed_menu_items():
    """Seed menu items from Solaria menu PDF"""
    
    # Get categories
    nasi_goreng = Category.query.filter_by(name='Nasi Goreng').first()
    mie = Category.query.filter_by(name='Mie').first()
    kwetiau = Category.query.filter_by(name='Kwetiau').first()
    menu_lain = Category.query.filter_by(name='Menu Lain').first()
    paket = Category.query.filter_by(name='Paket').first()
    snack = Category.query.filter_by(name='Snack').first()
    minuman = Category.query.filter_by(name='Minuman').first()
    
    menu_items_data = [
        # Nasi Goreng
        {'code': '111', 'name': 'Nasi Goreng Mlarat', 'price': 20000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '121', 'name': 'Nasi Goreng Spesial', 'price': 22000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '131', 'name': 'Nasi Goreng Cabe Ijo', 'price': 22000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=400&h=300&fit=crop'},
        {'code': '141', 'name': 'Nasi Goreng Sosis', 'price': 23000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1596560548464-f010549b84d7?w=400&h=300&fit=crop'},
        {'code': '151', 'name': 'Nasi Goreng Modern Warno', 'price': 24000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '161', 'name': 'Nasi Goreng Terimaskenthir', 'price': 25000, 'category': nasi_goreng, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '171', 'name': 'Nasi Goreng Pete', 'price': 25000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1569058242253-92a9c755a0ec?w=400&h=300&fit=crop'},
        {'code': '181', 'name': 'Nasi Goreng Seafood', 'price': 28000, 'category': nasi_goreng, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        
        # Mie
        {'code': '212', 'name': 'Mie Goreng Ayam', 'price': 22000, 'category': mie, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '222', 'name': 'Mie Siram Ayam', 'price': 22000, 'category': mie, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1555126634-323283e090fa?w=400&h=300&fit=crop'},
        {'code': '232', 'name': 'Mie Goreng Seafood', 'price': 28000, 'category': mie, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '242', 'name': 'Mie Siram Seafood', 'price': 28000, 'category': mie, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=400&h=300&fit=crop'},
        
        # Kwetiau
        {'code': '414', 'name': 'Kwetiau Ayam Goreng', 'price': 25000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '424', 'name': 'Kwetiau Ayam Siram', 'price': 25000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        {'code': '434', 'name': 'Kwetiau Seafood Goreng', 'price': 28000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '444', 'name': 'Kwetiau Seafood Siram', 'price': 28000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=400&h=300&fit=crop'},
        {'code': '454', 'name': 'Kwetiau Sapi Goreng', 'price': 30000, 'category': kwetiau, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '464', 'name': 'Kwetiau Sapi Siram', 'price': 30000, 'category': kwetiau, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=400&h=300&fit=crop'},
        
        # Menu Lain
        {'code': '515', 'name': 'Cap Cay Goreng Ayam', 'price': 23000, 'category': menu_lain, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '525', 'name': 'Cap Cay Goreng Seafood', 'price': 28000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=400&h=300&fit=crop'},
        {'code': '535', 'name': 'Sapo Tahu Ayam', 'price': 27000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=400&h=300&fit=crop'},
        {'code': '545', 'name': 'Sapo Tahu Seafood', 'price': 30000, 'category': menu_lain, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1512058564366-18510be2db19?w=400&h=300&fit=crop'},
        {'code': '555', 'name': 'Nasi Putih', 'price': 5000, 'category': menu_lain, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1516684732162-798a0062be99?w=400&h=300&fit=crop'},
        
        # Snack
        {'code': '313', 'name': 'Fish Cake', 'price': 12000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '323', 'name': 'Kentang Goreng', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        {'code': '333', 'name': 'Otak Otak', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '343', 'name': 'Sosis Goreng', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1612874742237-6526221588e3?w=400&h=300&fit=crop'},
        {'code': '353', 'name': 'Sosis Bakar', 'price': 15000, 'category': snack, 'has_spicy': False, 'popular': False, 'image': 'https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=400&h=300&fit=crop'},
        {'code': '363', 'name': 'Mix OTP', 'price': 20000, 'category': snack, 'has_spicy': False, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        
        # Paket
        {'code': '616', 'name': 'Nasi Goreng Cabe Ijo + Teh', 'price': 25000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '626', 'name': 'Kwetiau Ayam Goreng + Teh', 'price': 28000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '636', 'name': 'Nasi Goreng Spesial + Lemon Tea', 'price': 33000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '646', 'name': 'Kwetiau Ayam Goreng + Thai Tea', 'price': 35000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1585032226651-759b368d7246?w=400&h=300&fit=crop'},
        {'code': '656', 'name': '2 Thai Tea + Kentang Goreng', 'price': 38000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1576107232684-1279f390859f?w=400&h=300&fit=crop'},
        {'code': '666', 'name': '2 Cappucino + Mix OTP', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=400&h=300&fit=crop'},
        {'code': '676', 'name': 'Nasi Goreng + Kwetiau Seafood + Blackcurant', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': True, 'image': 'https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=400&h=300&fit=crop'},
        {'code': '686', 'name': 'Nasi + Sapo Tahu Seafood + Lemonade', 'price': 45000, 'category': paket, 'has_spicy': True, 'popular': False, 'image': 'https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=400&h=300&fit=crop'},
        
        # Minuman
        {'code': '717', 'name': 'Teh Mlarat', 'price': 3000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1576092768241-dec231879fc3?w=400&h=300&fit=crop'},
        {'code': '727', 'name': 'Teh Manis', 'price': 5000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1576092768241-dec231879fc3?w=400&h=300&fit=crop'},
        {'code': '737', 'name': 'Air Mineral', 'price': 5000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': False, 'image': 'https://images.unsplash.com/photo-1523362628745-0c100150b504?w=400&h=300&fit=crop'},
        {'code': '747', 'name': 'Kopi Hitam', 'price': 6000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1514432324607-a09d9b4aefdd?w=400&h=300&fit=crop'},
        {'code': '757', 'name': 'Green Tea', 'price': 13000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1556679343-c7306c1976bc?w=400&h=300&fit=crop'},
        {'code': '767', 'name': 'Thai Tea', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1558857563-b371033873b8?w=400&h=300&fit=crop'},
        {'code': '777', 'name': 'Green Tea Milk', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1515823064-d6e0c04616a7?w=400&h=300&fit=crop'},
        {'code': '787', 'name': 'Milo', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1517578239113-b03992dcdd25?w=400&h=300&fit=crop'},
        {'code': '797', 'name': 'Lemon Tea', 'price': 12000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1556679343-c7306c1976bc?w=400&h=300&fit=crop'},
        {'code': '807', 'name': 'Cappucino', 'price': 18000, 'category': minuman, 'has_spicy': False, 'popular': True, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1509042239860-f550ce710b93?w=400&h=300&fit=crop'},
        {'code': '817', 'name': 'Lemonade', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1621263764928-df1444c5e859?w=400&h=300&fit=crop'},
        {'code': '827', 'name': 'Blackcurrant', 'price': 15000, 'category': minuman, 'has_spicy': False, 'popular': False, 'has_temp': True, 'image': 'https://images.unsplash.com/photo-1544145945-f90425340c7e?w=400&h=300&fit=crop'},
    ]
    
    for item_data in menu_items_data:
        existing_item = MenuItem.query.filter_by(code=item_data['code']).first()
        if not existing_item:
            menu_item = MenuItem(
                code=item_data['code'],
                name=item_data['name'],
                price=item_data['price'],
                category_id=item_data['category'].id if item_data['category'] else None,
                has_spicy_option=item_data.get('has_spicy', False),
                has_temperature_option=item_data.get('has_temp', False),
                is_popular=item_data.get('popular', False),
                image=item_data.get('image', ''),
                description=f"Menu {item_data['name']} yang lezat"
            )
            db.session.add(menu_item)
        else:
            # Update image URL if it changed
            existing_item.image = item_data.get('image', existing_item.image)
    
    db.session.commit()

# Generate QR Code for table
def generate_table_qr(table_number):
    app_url = app.config.get('APP_URL', 'http://localhost:8000')
    order_url = f"{app_url}/order/online/{table_number}"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(order_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to file
    qr_folder = app.config.get('QR_CODE_FOLDER', 'static/qrcodes')
    os.makedirs(qr_folder, exist_ok=True)
    qr_path = os.path.join(qr_folder, f"table_{table_number}.png")
    img.save(qr_path)
    
    # Also return base64 for display
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    
    return qr_path, img_str

# ==================== ROUTES ====================

# Public routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Akun Anda telah dinonaktifkan. Hubungi administrator.', 'danger')
                return render_template('auth/login.html')
            
            login_user(user, remember=remember)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            next_page = request.args.get('next')
            flash(f'Selamat datang, {user.full_name or user.username}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Username atau password salah!', 'danger')
    
    return render_template('auth/login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        
        # Validation
        if User.query.filter_by(username=username).first():
            flash('Username sudah digunakan!', 'danger')
            return render_template('auth/register.html')
        
        if User.query.filter_by(email=email).first():
            flash('Email sudah digunakan!', 'danger')
            return render_template('auth/register.html')
        
        if password != confirm_password:
            flash('Password tidak cocok!', 'danger')
            return render_template('auth/register.html')
        
        if len(password) < 6:
            flash('Password minimal 6 karakter!', 'danger')
            return render_template('auth/register.html')
        
        # Create user
        customer_role = Role.query.filter_by(name='customer').first()
        user = User(
            username=username,
            email=email,
            full_name=full_name
        )
        user.set_password(password)
        if customer_role:
            user.roles.append(customer_role)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registrasi berhasil! Silakan login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('auth/register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))

# Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    today = datetime.now().date()
    
    # Get only completed/paid orders today
    today_orders = Order.query.filter(
        db.func.date(Order.created_at) == today
    ).all()
    
    # Calculate income from paid orders only
    paid_orders_today = [o for o in today_orders if o.payment and o.payment.status == 'paid']
    total_income_today = sum(o.total for o in paid_orders_today)
    total_orders_today = len(paid_orders_today)
    
    # Get popular items from actual order statistics (last 30 days)
    # Query to find most ordered items
    popular_query = db.session.query(
        OrderItem.menu_item_id,
        db.func.sum(OrderItem.quantity).label('total_ordered')
    ).join(Order).filter(
        Order.created_at >= datetime.now() - timedelta(days=30),
        Order.status.in_(['completed', 'processing'])
    ).group_by(OrderItem.menu_item_id).order_by(
        db.func.sum(OrderItem.quantity).desc()
    ).limit(6).all()
    
    # Get menu items for popular items
    popular_item_ids = [item[0] for item in popular_query if item[0]]
    if popular_item_ids:
        popular_items = MenuItem.query.filter(MenuItem.id.in_(popular_item_ids)).all()
        # Sort by order count
        item_order = {item_id: idx for idx, (item_id, _) in enumerate(popular_query)}
        popular_items.sort(key=lambda x: item_order.get(x.id, 999))
    else:
        # Fallback to marked popular items if no orders yet
        popular_items = MenuItem.query.filter_by(is_popular=True).limit(6).all()
    
    # Get recent orders
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(10).all()
    
    # Get tables status - based on active orders
    tables = Table.query.filter_by(is_active=True).all()
    
    # Calculate occupied tables from current active orders
    active_orders = Order.query.filter(
        Order.status.in_(['pending', 'processing']),
        Order.table_id.isnot(None)
    ).all()
    occupied_table_ids = set(o.table_id for o in active_orders)
    
    # Update tables status dynamically
    for table in tables:
        if table.id in occupied_table_ids:
            table.status = 'occupied'
        else:
            table.status = 'available'
    
    return render_template('dashboard.html',
                         total_income_today=total_income_today,
                         total_orders_today=total_orders_today,
                         popular_items=popular_items,
                         recent_orders=recent_orders,
                         tables=tables,
                         now=datetime.now())

# POS (Kasir)
@app.route('/pos')
@login_required
@role_required('admin', 'manager', 'kasir')
def pos():
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    menu_items = MenuItem.query.filter_by(is_available=True).all()
    tables = Table.query.filter_by(is_active=True).all()
    
    return render_template('pos.html',
                         categories=categories,
                         menu_items=menu_items,
                         tables=tables,
                         now=datetime.now())

# Online Order (via QR code)
@app.route('/order/online/<table_number>')
def online_order(table_number):
    table = Table.query.filter_by(number=table_number).first()
    if not table:
        flash('Meja tidak ditemukan!', 'danger')
        return redirect(url_for('login'))
    
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    menu_items = MenuItem.query.filter_by(is_available=True).all()
    
    return render_template('online_order.html',
                         table=table,
                         categories=categories,
                         menu_items=menu_items,
                         now=datetime.now())

# API Routes
@app.route('/api/menu')
def api_get_menu():
    menu_items = MenuItem.query.filter_by(is_available=True).all()
    return jsonify([item.to_dict() for item in menu_items])

@app.route('/api/menu/category/<int:category_id>')
def api_get_menu_by_category(category_id):
    menu_items = MenuItem.query.filter_by(category_id=category_id, is_available=True).all()
    return jsonify([item.to_dict() for item in menu_items])

# ============================================
# CART API - Database-backed shopping cart
# ============================================

def get_or_create_cart():
    """Get current user's cart or create new one"""
    if current_user.is_authenticated:
        cart = Cart.query.filter_by(user_id=current_user.id).first()
        if not cart:
            cart = Cart(user_id=current_user.id)
            db.session.add(cart)
            db.session.commit()
    else:
        # For guest users, use session
        session_id = session.get('cart_session_id')
        if not session_id:
            import uuid
            session_id = str(uuid.uuid4())
            session['cart_session_id'] = session_id
        
        cart = Cart.query.filter_by(session_id=session_id).first()
        if not cart:
            cart = Cart(session_id=session_id)
            db.session.add(cart)
            db.session.commit()
    
    return cart

@app.route('/api/cart')
def api_get_cart():
    """Get current cart items"""
    cart = get_or_create_cart()
    return jsonify({'success': True, 'cart': cart.to_dict()})

@app.route('/api/cart/add', methods=['POST'])
def api_add_to_cart():
    """Add item to cart"""
    try:
        data = request.json
        menu_item_id = data.get('menu_item_id')
        quantity = data.get('quantity', 1)
        spice_level = data.get('spice_level')
        temperature = data.get('temperature')
        notes = data.get('notes', '')
        
        menu_item = db.session.get(MenuItem, menu_item_id)
        if not menu_item:
            return jsonify({'success': False, 'error': 'Menu item not found'}), 404
        
        cart = get_or_create_cart()
        
        # Check if same item with same options exists
        existing_item = CartItem.query.filter_by(
            cart_id=cart.id,
            menu_item_id=menu_item_id,
            spice_level=spice_level,
            temperature=temperature,
            notes=notes
        ).first()
        
        if existing_item:
            existing_item.quantity += quantity
            existing_item.update_subtotal()
        else:
            cart_item = CartItem(
                cart_id=cart.id,
                menu_item_id=menu_item_id,
                name=menu_item.name,
                price=menu_item.price,
                quantity=quantity,
                subtotal=menu_item.price * quantity,
                spice_level=spice_level,
                temperature=temperature,
                notes=notes
            )
            db.session.add(cart_item)
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/update/<int:item_id>', methods=['PUT'])
def api_update_cart_item(item_id):
    """Update cart item quantity"""
    try:
        data = request.json
        quantity = data.get('quantity', 1)
        
        cart = get_or_create_cart()
        cart_item = CartItem.query.filter_by(id=item_id, cart_id=cart.id).first()
        
        if not cart_item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        
        if quantity <= 0:
            db.session.delete(cart_item)
        else:
            cart_item.quantity = quantity
            cart_item.update_subtotal()
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/remove/<int:item_id>', methods=['DELETE'])
def api_remove_cart_item(item_id):
    """Remove item from cart"""
    try:
        cart = get_or_create_cart()
        cart_item = CartItem.query.filter_by(id=item_id, cart_id=cart.id).first()
        
        if not cart_item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        
        db.session.delete(cart_item)
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/clear', methods=['DELETE'])
def api_clear_cart():
    """Clear all items from cart"""
    try:
        cart = get_or_create_cart()
        CartItem.query.filter_by(cart_id=cart.id).delete()
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cart/settings', methods=['PUT'])
def api_update_cart_settings():
    """Update cart settings (table, order type, customer name)"""
    try:
        data = request.json
        cart = get_or_create_cart()
        
        if 'table_id' in data:
            cart.table_id = data['table_id'] if data['table_id'] else None
        if 'order_type' in data:
            cart.order_type = data['order_type']
        if 'customer_name' in data:
            cart.customer_name = data['customer_name']
        
        db.session.commit()
        return jsonify({'success': True, 'cart': cart.to_dict()})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================

@app.route('/api/order', methods=['POST'])
def api_create_order():
    try:
        data = request.json
        items = data.get('items', [])
        table_id = data.get('table_id')
        order_type = data.get('order_type', 'dine_in')
        customer_name = data.get('customer_name', '')
        notes = data.get('notes', '')
        payment_method = data.get('payment_method', 'cash')
        paid_amount = data.get('paid_amount', 0)
        
        if not items:
            return jsonify({'error': 'Keranjang kosong'}), 400
        
        # Generate order number
        order_number = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Create order
        order = Order(
            order_number=order_number,
            user_id=current_user.id if current_user.is_authenticated else None,
            table_id=table_id,
            customer_name=customer_name,
            order_type=order_type,
            notes=notes
        )
        db.session.add(order)
        db.session.flush()
        
        # Add order items
        for item in items:
            menu_item = db.session.get(MenuItem, item.get('menu_item_id') or item.get('id'))
            if menu_item:
                order_item = OrderItem(
                    order_id=order.id,
                    menu_item_id=menu_item.id,
                    name=menu_item.name,
                    price=menu_item.price,
                    quantity=item.get('quantity', 1),
                    subtotal=menu_item.price * item.get('quantity', 1),
                    spice_level=item.get('spice_level'),
                    temperature=item.get('temperature'),
                    notes=item.get('notes')
                )
                db.session.add(order_item)
        
        # Calculate totals
        order.calculate_totals()
        
        # Create payment record
        is_cash_paid = payment_method == 'cash' and paid_amount >= order.total
        payment = Payment(
            order_id=order.id,
            payment_method=payment_method,
            amount=order.total,
            paid_amount=paid_amount if payment_method == 'cash' else 0,
            change_amount=max(0, paid_amount - order.total) if payment_method == 'cash' else 0,
            status='paid' if is_cash_paid else 'pending'
        )
        
        if is_cash_paid:
            payment.paid_at = datetime.utcnow()
            order.status = 'processing'
        
        db.session.add(payment)
        db.session.commit()
        
        # Update table status
        if table_id:
            table = db.session.get(Table, table_id)
            if table:
                table.status = 'occupied'
                db.session.commit()
        
        # For online payment (Midtrans), generate Snap token
        if payment_method == 'online':
            midtrans_order_id = f"DTO-{order.order_number}"
            payment.midtrans_order_id = midtrans_order_id
            payment.payment_method = 'midtrans'
            
            # Generate Midtrans Snap token using Snap API
            snap_token = generate_midtrans_snap_token(order, midtrans_order_id)
            
            if snap_token:
                payment.snap_token = snap_token
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'order': order.to_dict(),
                    'payment_method': 'online',
                    'snap_token': snap_token,
                    'midtrans_client_key': app.config.get('MIDTRANS_CLIENT_KEY'),
                    'message': 'Pesanan dibuat, silakan lakukan pembayaran.'
                })
            else:
                # Fallback: still create order but mark as pending
                db.session.commit()
                return jsonify({
                    'success': True,
                    'order': order.to_dict(),
                    'payment_method': 'online',
                    'snap_token': None,
                    'error_payment': 'Gagal menghubungi payment gateway, silakan coba lagi.',
                    'message': 'Pesanan dibuat dengan status pending.'
                })
        
        # Clear cart after successful order
        if current_user.is_authenticated:
            cart = Cart.query.filter_by(user_id=current_user.id).first()
        else:
            session_id = session.get('cart_session_id')
            cart = Cart.query.filter_by(session_id=session_id).first() if session_id else None
        
        if cart:
            CartItem.query.filter_by(cart_id=cart.id).delete()
            db.session.commit()
        
        return jsonify({
            'success': True,
            'order': order.to_dict(),
            'payment_method': 'cash',
            'message': 'Pesanan berhasil dibuat!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def generate_midtrans_snap_token(order, midtrans_order_id):
    """Generate Midtrans Snap token for payment"""
    import requests
    import base64
    
    server_key = app.config.get('MIDTRANS_SERVER_KEY', '')
    is_production = app.config.get('MIDTRANS_IS_PRODUCTION', False)
    
    # Determine API URL
    if is_production:
        snap_url = 'https://app.midtrans.com/snap/v1/transactions'
    else:
        snap_url = 'https://app.sandbox.midtrans.com/snap/v1/transactions'
    
    # Prepare transaction details
    transaction_details = {
        'order_id': midtrans_order_id,
        'gross_amount': int(order.total)
    }
    
    # Prepare item details
    item_details = []
    for item in order.items:
        item_details.append({
            'id': str(item.menu_item_id),
            'price': int(item.price),
            'quantity': item.quantity,
            'name': item.name[:50]  # Midtrans limits name to 50 chars
        })
    
    # Add tax as item
    if order.tax > 0:
        item_details.append({
            'id': 'TAX',
            'price': int(order.tax),
            'quantity': 1,
            'name': 'Pajak (10%)'
        })
    
    # Customer details
    customer_details = {
        'first_name': order.customer_name or 'Customer',
        'email': 'customer@kasir.local'
    }
    
    if current_user.is_authenticated:
        customer_details['first_name'] = current_user.full_name or current_user.username
        customer_details['email'] = current_user.email or 'customer@kasir.local'
    
    # Build request payload
    payload = {
        'transaction_details': transaction_details,
        'item_details': item_details,
        'customer_details': customer_details
    }
    
    # Create authorization header
    auth_string = base64.b64encode(f"{server_key}:".encode()).decode()
    headers = {
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'Authorization': f'Basic {auth_string}'
    }
    
    try:
        response = requests.post(snap_url, json=payload, headers=headers, timeout=30)
        if response.status_code == 201:
            data = response.json()
            return data.get('token')
        else:
            print(f"Midtrans error: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Midtrans connection error: {e}")
        return None

@app.route('/api/order/<int:order_id>/status', methods=['PUT'])
@login_required
def api_update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.json
    
    order.status = data.get('status', order.status)
    
    if order.status == 'completed' and order.table:
        order.table.status = 'available'
    
    db.session.commit()
    
    return jsonify({'success': True, 'order': order.to_dict()})

@app.route('/api/payment/midtrans', methods=['POST'])
def api_create_midtrans_payment():
    try:
        data = request.json
        order_id = data.get('order_id')
        
        order = Order.query.get_or_404(order_id)
        
        # Create Midtrans transaction (simplified - in production use midtransclient)
        midtrans_order_id = f"DTO-{order.order_number}"
        
        # Update payment
        if order.payment:
            order.payment.payment_method = 'midtrans'
            order.payment.midtrans_order_id = midtrans_order_id
            order.payment.status = 'pending'
            # In production, generate actual Snap token here
            order.payment.payment_url = f"https://app.sandbox.midtrans.com/snap/v2/vtweb/{midtrans_order_id}"
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'payment_url': order.payment.payment_url,
            'order_id': midtrans_order_id
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/payment/midtrans/callback', methods=['POST'])
def api_midtrans_callback():
    try:
        data = request.json
        order_id = data.get('order_id')
        transaction_status = data.get('transaction_status')
        
        # Find payment by midtrans_order_id
        payment = Payment.query.filter_by(midtrans_order_id=order_id).first()
        
        if payment:
            payment.midtrans_status = transaction_status
            
            if transaction_status in ['capture', 'settlement']:
                payment.status = 'paid'
                payment.paid_at = datetime.utcnow()
                payment.order.status = 'processing'
            elif transaction_status in ['deny', 'cancel', 'expire']:
                payment.status = 'failed'
            
            db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Statistics API
@app.route('/api/stats')
@login_required
def api_get_stats():
    today = datetime.now().date()
    
    # Today's stats
    today_orders = Order.query.filter(
        db.func.date(Order.created_at) == today
    ).all()
    
    total_income = sum(o.total for o in today_orders if o.payment and o.payment.status == 'paid')
    total_orders = len(today_orders)
    
    # Popular items today
    popular_items = {}
    for order in today_orders:
        for item in order.items:
            if item.name not in popular_items:
                popular_items[item.name] = 0
            popular_items[item.name] += item.quantity
    
    most_popular = sorted(popular_items.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return jsonify({
        'total_income': total_income,
        'total_orders': total_orders,
        'most_popular': most_popular,
        'average_transaction': total_income / total_orders if total_orders > 0 else 0
    })

# Profile
@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    full_name = request.form.get('full_name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    
    # Check email uniqueness
    if email != current_user.email:
        if User.query.filter_by(email=email).first():
            flash('Email sudah digunakan!', 'danger')
            return redirect(url_for('profile'))
    
    current_user.full_name = full_name
    current_user.email = email
    current_user.phone = phone
    
    db.session.commit()
    flash('Profil berhasil diperbarui!', 'success')
    return redirect(url_for('profile'))

@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not current_user.check_password(current_password):
        flash('Password saat ini salah!', 'danger')
        return redirect(url_for('profile'))
    
    if new_password != confirm_password:
        flash('Password baru tidak cocok!', 'danger')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('Password minimal 6 karakter!', 'danger')
        return redirect(url_for('profile'))
    
    current_user.set_password(new_password)
    db.session.commit()
    flash('Password berhasil diubah!', 'success')
    return redirect(url_for('profile'))

# Admin routes
@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    users = User.query.all()
    roles = Role.query.all()
    return render_template('admin/users.html', users=users, roles=roles)

@app.route('/admin/users/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    role_id = request.form.get('role_id')
    
    if User.query.filter_by(username=username).first():
        flash('Username sudah digunakan!', 'danger')
        return redirect(url_for('admin_users'))
    
    user = User(
        username=username,
        email=email,
        full_name=full_name
    )
    user.set_password(password)
    
    if role_id:
        role = db.session.get(Role, role_id)
        if role:
            user.roles.append(role)
    
    db.session.add(user)
    db.session.commit()
    
    flash('User berhasil dibuat!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'diaktifkan' if user.is_active else 'dinonaktifkan'
    flash(f'User {user.username} berhasil {status}!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/menu')
@login_required
@role_required('admin', 'manager')
def admin_menu():
    categories = Category.query.order_by(Category.order).all()
    menu_items = MenuItem.query.all()
    return render_template('admin/menu.html', categories=categories, menu_items=menu_items)

@app.route('/admin/menu/create', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def admin_create_menu():
    code = request.form.get('code')
    name = request.form.get('name')
    price = int(request.form.get('price', 0))
    category_id = request.form.get('category_id')
    description = request.form.get('description')
    is_popular = request.form.get('is_popular') == 'on'
    has_spicy_option = request.form.get('has_spicy_option') == 'on'
    has_temperature_option = request.form.get('has_temperature_option') == 'on'
    image = request.form.get('image')
    
    menu_item = MenuItem(
        code=code,
        name=name,
        price=price,
        category_id=category_id,
        description=description,
        is_popular=is_popular,
        has_spicy_option=has_spicy_option,
        has_temperature_option=has_temperature_option,
        image=image
    )
    
    db.session.add(menu_item)
    db.session.commit()
    
    flash('Menu berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_menu'))

@app.route('/admin/tables')
@login_required
@role_required('admin', 'manager')
def admin_tables():
    tables = Table.query.all()
    return render_template('admin/tables.html', tables=tables)

@app.route('/admin/tables/<int:table_id>/qr')
@login_required
@role_required('admin', 'manager')
def admin_table_qr(table_id):
    table = Table.query.get_or_404(table_id)
    qr_path, qr_base64 = generate_table_qr(table.number)
    table.qr_code = qr_path
    db.session.commit()
    
    return jsonify({
        'success': True,
        'qr_code': f"data:image/png;base64,{qr_base64}",
        'table_number': table.number
    })

# Reports
@app.route('/reports')
@login_required
@role_required('admin', 'manager')
def reports():
    return render_template('reports.html')

@app.route('/reports/income')
@login_required
@role_required('admin', 'manager')
def income_report():
    # Get date range from query params
    start_date_str = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    # Parse dates properly
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
    
    orders = Order.query.filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    total_income = sum(o.total for o in paid_orders)
    total_orders = len(paid_orders)
    
    # Group by date
    daily_income = {}
    for order in paid_orders:
        date_str = order.created_at.strftime('%Y-%m-%d')
        if date_str not in daily_income:
            daily_income[date_str] = {'income': 0, 'orders': 0}
        daily_income[date_str]['income'] += order.total
        daily_income[date_str]['orders'] += 1
    
    return render_template('reports/income.html',
                         orders=paid_orders,
                         total_income=total_income,
                         total_orders=total_orders,
                         daily_income=daily_income,
                         start_date=start_date_str,
                         end_date=end_date_str)

@app.route('/reports/export/pdf')
@login_required
@role_required('admin', 'manager')
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as PDFTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    orders = Order.query.filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date + ' 23:59:59'
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph(f"Laporan Penjualan", styles['Heading1']))
    elements.append(Paragraph(f"Periode: {start_date} - {end_date}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Tanggal', 'Order ID', 'Total']]
    for i, order in enumerate(paid_orders, 1):
        data.append([
            str(i),
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.order_number,
            f"Rp {order.total:,}".replace(',', '.')
        ])
    
    # Total row
    total = sum(o.total for o in paid_orders)
    data.append(['', '', 'TOTAL', f"Rp {total:,}".replace(',', '.')])
    
    table = PDFTable(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'laporan_{start_date}_{end_date}.pdf',
        mimetype='application/pdf'
    )

@app.route('/reports/export/excel')
@login_required
@role_required('admin', 'manager')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    orders = Order.query.filter(
        Order.created_at >= start_date,
        Order.created_at <= end_date + ' 23:59:59'
    ).all()
    
    paid_orders = [o for o in orders if o.payment and o.payment.status == 'paid']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    
    # Header
    ws['A1'] = 'Laporan Penjualan'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Periode: {start_date} - {end_date}'
    
    # Column headers
    headers = ['No', 'Tanggal', 'Order ID', 'Customer', 'Subtotal', 'Tax', 'Total', 'Payment Method']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, order in enumerate(paid_orders, 5):
        ws.cell(row=row, column=1, value=row-4)
        ws.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=order.order_number)
        ws.cell(row=row, column=4, value=order.customer_name or '-')
        ws.cell(row=row, column=5, value=order.subtotal)
        ws.cell(row=row, column=6, value=order.tax)
        ws.cell(row=row, column=7, value=order.total)
        ws.cell(row=row, column=8, value=order.payment.payment_method if order.payment else '-')
    
    # Total
    total_row = len(paid_orders) + 5
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(o.total for o in paid_orders)).font = Font(bold=True)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'laporan_{start_date}_{end_date}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Order management
@app.route('/orders')
@login_required
@role_required('admin', 'manager', 'kasir')
def orders():
    status_filter = request.args.get('status', 'all')
    
    query = Order.query.order_by(Order.created_at.desc())
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    orders = query.limit(100).all()
    
    return render_template('orders.html', orders=orders, status_filter=status_filter)

# Error handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('errors/500.html'), 500

if __name__ == '__main__':
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    os.makedirs('static/qrcodes', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    os.makedirs('uploads', exist_ok=True)
    
    # Initialize database
    init_db()
    
    print("""
      KASIR MODERN - FULL FEATURES
    ====================================
     FITUR:
    1.  Login & Register
    2.  Role & Permission
    3.  Pesanan Manual & Online (QR Code)
    4.  Profile & Logout
    5.  Payment Gateway (Midtrans)
    6.  Spice Level & Hot/Cold Options
    7.  Statistics & Reports (PDF & Excel)
    8.  Admin Management
    9.  Income Management
    10.  Menu dari PDF Solaria
    
     Modern UI dengan Tailwind CSS
     Glassmorphism Design
     Secure Authentication
    
     Default Login:
       Admin: admin / admin123
       Kasir: kasir / kasir123
    
     Server: http://localhost:8000
    """)
    
    # Use debug mode only in development (controlled by environment variable)
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=8000, use_reloader=debug_mode)
